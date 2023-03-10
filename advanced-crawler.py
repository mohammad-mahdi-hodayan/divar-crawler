'''https://api.divar.ir/v8/postcontact/web/contact_info/wY_2uZU7'''
from advanced_config import *
if (test_val) :
    print("config file loaded.")
else : 
    print("error : we couldn't find config file !")
from selenium import webdriver
import json
import time
import openpyxl 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
wb = Workbook(write_only = True)
ws = wb.create_sheet()
options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_argument("user-data-dir="+browser_profile) 
driver = webdriver.Chrome(options=options)
headers = ['title','subtitle','phone','table','description']
ws.append(headers) 
i = 0
with open(input_jsonfile, 'r', encoding="utf8") as json_file:
    data = json.load(json_file)
    for tx in data :
        i = i + 1 
        if i < crawl_starts_from :
            continue
        if i > crawl_limit :
            break
        dataname = "t"+str(i)
        driver.get(data[dataname]["link"])
        if (driver.find_elements_by_css_selector('.not-found-message .title')) :
            print("404 error")
            continue
        time.sleep(wait_before_click)
        driver.find_element_by_css_selector('button.post-actions__get-contact').click()
        print("item : " + str(i))
        time.sleep(wait_after_click)
        # selectors and elements :
        subtitle = ""
        title = ""
        phone = ""
        Frows = ""
        subtitle_elements = driver.find_elements_by_css_selector(".kt-page-title__subtitle")
        for sub in subtitle_elements : 
            subtitle = sub.get_attribute('textContent')
        title_elements = driver.find_elements_by_css_selector(".kt-page-title__title--responsive-sized")
        for tit in title_elements : 
            title = tit.get_attribute('textContent')
        phone_elements = driver.find_elements_by_css_selector(".copy-row a.kt-unexpandable-row__action")
        for pel in phone_elements : 
            phone = pel.get_attribute('href')
            phone = phone.replace("tel:","")
        Frows = ""
        row_elements = driver.find_elements_by_css_selector(".post-page__section--padded .kt-base-row.kt-unexpandable-row")
        for row in row_elements : 
            dataname = row.find_element_by_css_selector(".kt-base-row__start").get_attribute('textContent')
            dataval = row.find_element_by_css_selector(".kt-base-row__end").get_attribute('textContent')
            row_text = str(dataname) +" : "+ str(dataval) 
            Frows = Frows + str(row_text) + "\n"
        description_elements = driver.find_elements_by_css_selector(".kt-description-row__text")
        for description_box in description_elements : 
            description = description_box.get_attribute('textContent')
        ws.append([title,subtitle,phone,Frows,description])
print("crawl finished. generating xcel file...")
wb.save(output_xlsxfile)
driver.close()
time.sleep(1)
exit()